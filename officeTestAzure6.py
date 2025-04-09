import os
import logging
import re
from azure.storage.blob import BlobServiceClient
import zipfile
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from typing import List, Dict, Optional
from datetime import datetime
import calendar
from concurrent.futures import ThreadPoolExecutor
import multiprocessing

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('file_processor.log'),
        logging.StreamHandler()
    ]
)

# Azure Connection String
AZURE_CONNECTION_STRING = os.getenv("AZURE_CONNECTION_STRING")
if not AZURE_CONNECTION_STRING:
    raise ValueError("Azure connection string not found in environment variables")

# Complete Rename Mapping
RENAME_MAPPING = {
    "CT1 IT.xlsx": "IT_CT1.xlsx",
    "CT1 Mains.xlsx": "Electricity_Mains_CT1.xlsx",
    "CT1 Diesel.xlsx": "Diesel_CT1.xlsx",
    "CT1 Solar.xlsx": "Solar_CT1.xlsx",
    "CT1 PUE report.xlsx": "PUE_CT1.xlsx",
    "CT IT loads.xls": "IT_Loads_CT1.xlsx",
    "CT2 IT.xlsx": "IT_CT2.xlsx",
    "CT2 Mains.xlsx": "Electricity_Mains_CT2.xlsx",
    "CT2 Diesel.xlsx": "Diesel_CT2.xlsx",
    "CT2 Solar.xlsx": "Solar_CT2.xlsx",
    "CT2 PUE report.xlsx": "PUE_CT2.xlsx",
    "ct2 IT load.xls": "IT_Loads_CT2.xlsx",
    "CT2 IT loads.xls": "IT_Loads_CT2.xlsx",
    "DB1 IT.xlsx": "IT_DB1.xlsx",
    "DB1 Mains.xlsx": "Electricity_Mains_DB1.xlsx",
    "DB1 Diesel.xlsx": "Diesel_DB1.xlsx",
    "DB1 Solar.xlsx": "Solar_DB1.xlsx",
    "DB1 PUE report.xlsx": "PUE_DB1.xlsx",
    "DB1 IT load.xls": "IT_Loads_DB1.xlsx",
    "JB-E IT.xlsx": "IT_JB-E.xlsx",
    "JB-E Mains.xlsx": "Electricity_Mains_JB-E.xlsx",
    "JB-E Diesel.xlsx": "Diesel_JB-E.xlsx",
    "JB-E Solar.xlsx": "Solar_JB-E.xlsx",
    "JB-E PUE report.xlsx": "PUE_JB-E.xlsx",
    "JBE IT load.xls": "IT_Loads_JB-E.xlsx",
    "JB-W IT.xlsx": "IT_JB-W.xlsx",
    "JB-W Mains.xlsx": "Electricity_Mains_JB-W.xlsx",
    "JB-W Diesel.xlsx": "Diesel_JB-W.xlsx",
    "JB-W Solar.xlsx": "Solar_JB-W.xlsx",
    "JB-W PUE report.xlsx": "PUE_JB-W.xlsx",
    "JB-W PUE Report.xlsx": "PUE_JB-W.xlsx",
    "JBW IT load.xls": "IT_Loads_JB-W.xlsx",
    "JBW IT total.xls": "IT_Loads_JB-W.xlsx",
    "JB2 IT.xlsx": "IT_JB2.xlsx",
    "JB2 Mains.xlsx": "Electricity_Mains_JB2.xlsx",
    "JB2 Diesel.xlsx": "Diesel_JB2.xlsx",
    "JB2 Solar.xlsx": "Solar_JB2.xlsx",
    "JB2 PUE report.xlsx": "PUE_JB2.xlsx",
    "JB2 IT loads.xls": "IT_Loads_JB2.xlsx",
    "JB3 IT.xlsx": "IT_JB3.xlsx",
    "JB3 Mains.xlsx": "Electricity_Mains_JB3.xlsx",
    "JB3 Diesel.xlsx": "Diesel_JB3.xlsx",
    "JB3 Solar.xlsx": "Solar_JB3.xlsx",
    "JB3 PUE report.xlsx": "PUE_JB3.xlsx",
    "JB3 IT loads.xls": "IT_Loads_JB3.xlsx",
    "JB4 IT.xlsx": "IT_JB4.xlsx",
    "JB4 Mains.xlsx": "Electricity_Mains_JB4.xlsx",
    "JB4 Diesel.xlsx": "Diesel_JB4.xlsx",
    "JB4 Solar.xlsx": "Solar_JB4.xlsx",
    "JB4 PUE report.xlsx": "PUE_JB4.xlsx",
    "JB4 PUE.xlsx": "PUE_JB4.xlsx",
    "JB4 IT Total (3).xls": "IT_Loads_JB4.xlsx",
    "JB4 IT loads.xls": "IT_Loads_JB4.xlsx",
    "TDE-IM-FT-028 - CT1 Refrigerant Use.xlsx": "Refrigerants_CT1.xlsx",
    "TDE-IM-FT-028 - CT2 Refrigerant Use.xlsx": "Refrigerants_CT2.xlsx",
    "TDE-IM-FT-028 - DB1 Refrigerant Use.xlsx": "Refrigerants_DB1.xlsx",
    "TDE-IM-FT-028 - JB1 Refrigerant Use.xlsx": "Refrigerants_JB1.xlsx",
    "TDE-IM-FT-028 - JB2 Refrigerant Use.xlsx": "Refrigerants_JB2.xlsx",
    "TDE-IM-FT-028 - JB3 Refrigerant Use.xlsx": "Refrigerants_JB3.xlsx",
    "TDE-IM-FT-028 - JB4 Refrigerant Use.xlsx": "Refrigerants_JB4.xlsx",
    "MIS REPORT - TERACO.xlsx": "CarRental_Consolidated.xlsx",
}

# Dynamic renaming patterns
DYNAMIC_RENAME_PATTERNS = {
    'water': re.compile(r"^.*WaterMeterRecon_Construction.*\.xlsx$", re.IGNORECASE),
    'carrental': re.compile(r"^.*MIS REPORT - TERACO.*\.xlsx$", re.IGNORECASE)
}

# Energy file prefixes
ENERGY_FILE_PREFIXES = ["Electricity_Mains_", "Diesel_", "Solar_", "PUE_", "IT_"]

# Water file tabs
WATER_FILE_TABS = ['JB1E', 'JB1W', 'JB2', 'JB3', 'JB4', 'CT1', 'CT2', 'DB1', 'Consumption Calcs']

def rename_worksheet_tab(file_path: str, new_tab_name: str) -> None:
    """Rename the first worksheet tab in an Excel file."""
    try:
        workbook = load_workbook(file_path, read_only=False, keep_vba=False)
        if workbook.sheetnames:
            old_tab = workbook.sheetnames[0]
            if old_tab != new_tab_name:
                workbook[old_tab].title = new_tab_name
                workbook.save(file_path, optimized_write=True)
    except Exception as e:
        logging.warning(f"Couldn't rename worksheet in {file_path}: {e}")

def copy_and_rename_carrental_tab(file_path: str) -> None:
    """Copy and rename the date range tab in CarRental file to 'PBI'"""
    try:
        if not os.path.exists(file_path) or "carrental_consolidated" not in file_path.lower():
            return

        workbook = load_workbook(file_path)
        source_tab = None
        
        # Find the tab that starts with date pattern
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("01 Jan -"):
                source_tab = sheet_name
                break
        
        if source_tab:
            # Create copy of the sheet
            source_sheet = workbook[source_tab]
            new_sheet = workbook.copy_worksheet(source_sheet)
            new_sheet.title = "PBI"
            workbook.save(file_path)
            logging.info(f"Copied tab '{source_tab}' to 'PBI' in {os.path.basename(file_path)}")
        else:
            logging.warning(f"No tab starting with '01 Jan -' found in {os.path.basename(file_path)}")
            
    except Exception as e:
        logging.error(f"Error processing CarRental tab in {file_path}: {e}")

def normalize_name(name: str) -> str:
    """Normalize file names for comparison."""
    return re.sub(r'\W+', '', name).lower()

def convert_to_excel(file_path: str) -> Optional[str]:
    """Convert files to Excel (.xlsx) format."""
    try:
        if file_path.endswith('.xls'):
            df = pd.read_excel(file_path, engine='xlrd')
        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            return None
            
        excel_path = file_path.rsplit('.', 1)[0] + '.xlsx'
        df.to_excel(excel_path, index=False, engine='openpyxl')
        os.remove(file_path)
        return excel_path
    except Exception as e:
        logging.warning(f"Couldn't convert {file_path}: {e}")
        return None

def rename_files(download_folder: str, rename_mapping: Dict[str, str]) -> None:
    """Rename files with pattern matching and worksheet tab updates."""
    try:
        logging.info("Starting file renaming process...")
        renamed_count = 0
        
        for root, _, files in os.walk(download_folder):
            for file in files:
                try:
                    file_path = os.path.join(root, file)
                    
                    # Convert to Excel if needed
                    if file_path.endswith(('.xls', '.csv')):
                        file_path = convert_to_excel(file_path) or file_path
                    
                    file_name = os.path.basename(file_path)
                    normalized_name = normalize_name(file_name)
                    renamed = False
                    
                    # Check exact mappings first
                    for original, new_name in rename_mapping.items():
                        if normalized_name == normalize_name(original):
                            new_path = os.path.join(root, new_name)
                            os.rename(file_path, new_path)
                            renamed_count += 1
                            logging.info(f"Renamed: {file} → {new_name}")
                            
                            # Special handling for CarRental files
                            if "carrental_consolidated" in new_name.lower():
                                rename_worksheet_tab(new_path, "CarRental_Consolidated")
                                copy_and_rename_carrental_tab(new_path)
                            renamed = True
                            break
                    
                    if not renamed:
                        # Check dynamic patterns
                        if DYNAMIC_RENAME_PATTERNS['water'].match(file_name):
                            new_name = "Water_Consolidated.xlsx"
                            new_path = os.path.join(root, new_name)
                            os.rename(file_path, new_path)
                            renamed_count += 1
                            logging.info(f"Renamed (water): {file} → {new_name}")
                            
                        elif DYNAMIC_RENAME_PATTERNS['carrental'].match(file_name):
                            new_name = "CarRental_Consolidated.xlsx"
                            new_path = os.path.join(root, new_name)
                            os.rename(file_path, new_path)
                            rename_worksheet_tab(new_path, "CarRental_Consolidated")
                            copy_and_rename_carrental_tab(new_path)
                            renamed_count += 1
                            logging.info(f"Renamed (carrental): {file} → {new_name}")
                            
                except Exception as e:
                    logging.warning(f"Error processing {file}: {e}")
                    continue
                    
        logging.info(f"Renaming complete. {renamed_count} files processed.")
    except Exception as e:
        logging.error(f"Error during file renaming: {e}")
        raise

def connect_to_azure() -> BlobServiceClient:
    """Connect to Azure Blob Storage with optimized settings"""
    return BlobServiceClient.from_connection_string(
        AZURE_CONNECTION_STRING,
        max_single_get_size=4*1024*1024,
        max_chunk_get_size=4*1024*1024
    )

def download_blob(container_client, blob_name: str, dest_path: str) -> None:
    """Download a blob with path creation"""
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, "wb") as f:
        container_client.get_blob_client(blob_name).download_blob().readinto(f)

def unzip_file(file_path: str, extract_to: str) -> None:
    """Unzip downloaded files"""
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)

def remove_last_row_for_energy_files(download_folder: str) -> None:
    """Remove last row from energy files"""
    try:
        for root, _, files in os.walk(download_folder):
            for file in files:
                if not file.endswith('.xlsx'):
                    continue
                    
                file_path = os.path.join(root, file)
                normalized_name = normalize_name(file)
                
                if any(normalized_name.startswith(normalize_name(prefix)) for prefix in ENERGY_FILE_PREFIXES):
                    try:
                        workbook = load_workbook(file_path)
                        sheet = workbook.active
                        last_row = None
                        for row in sheet.iter_rows(min_col=2, max_col=2):
                            if row[0].value is not None:
                                last_row = row[0].row
                        if last_row:
                            sheet.delete_rows(last_row)
                            workbook.save(file_path)
                    except Exception as e:
                        logging.warning(f"Couldn't process {file}: {e}")
    except Exception as e:
        logging.error(f"Error processing energy files: {e}")
        raise

def zero_out_cells_in_next_month(file_path: str, reporting_month: str) -> None:
    """Zero out cells in next month column for Water files"""
    try:
        month, year = parse_month_year(reporting_month)
        next_month, next_year = get_next_month(month, year)
        expected_next_header = format_month_year(next_month, next_year)
        
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        modified = False

        for sheet_name in WATER_FILE_TABS:
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]
            reporting_col = None
            
            for col_idx, cell in enumerate(sheet[2], 1):
                cell_value = str(cell.value).strip().lower() if cell.value else ""
                if isinstance(cell.value, datetime):
                    cell_value = cell.value.strftime("%b-%y").lower()
                if cell_value == reporting_month.strip().lower():
                    reporting_col = col_idx
                    break

            if reporting_col and reporting_col + 1 <= sheet.max_column:
                next_col = reporting_col + 1
                for row in sheet.iter_rows(min_row=3, min_col=next_col, max_col=next_col):
                    if not isinstance(row[0].value, datetime):
                        row[0].value = 0
                        modified = True

        if modified:
            workbook.save(file_path)
    except Exception as e:
        logging.error(f"Error processing {file_path}: {e}")
        raise

def parse_month_year(month_year_str: str) -> tuple:
    """Parse month-year string into (month, year) tuple"""
    try:
        parts = re.split(r'[-_]', month_year_str.strip())
        if len(parts) != 2:
            raise ValueError("Invalid format")
        month_str, year_str = parts
        month = None
        for fmt in ('%b', '%B'):
            try:
                month = datetime.strptime(month_str, fmt).month
                break
            except ValueError:
                continue
        if not month:
            raise ValueError("Invalid month")
        year = 2000 + int(year_str) if len(year_str) == 2 else int(year_str)
        return (month, year)
    except Exception as e:
        raise ValueError(f"Could not parse month-year string '{month_year_str}': {e}")

def get_next_month(month: int, year: int) -> tuple:
    """Get next month and year"""
    return (1, year + 1) if month == 12 else (month + 1, year)

def format_month_year(month: int, year: int) -> str:
    """Format month and year as 'MMM-YY'"""
    return f"{calendar.month_abbr[month]}-{str(year)[-2:]}"

def cleanup_temp_files(download_folder: str) -> None:
    """Clean up temporary files"""
    for root, _, files in os.walk(download_folder):
        for file in files:
            if file.endswith(('.csv', '.zip')):
                try:
                    os.remove(os.path.join(root, file))
                except Exception:
                    pass

def display_files(download_folder: str) -> None:
    """Display all files in download folder"""
    print("\nFiles in download folder:")
    for root, _, files in os.walk(download_folder):
        for file in files:
            print(f"  • {os.path.join(root, file)}")

def find_water_file(download_folder: str) -> Optional[str]:
    """Find Water_Consolidated.xlsx in download folder"""
    for root, _, files in os.walk(download_folder):
        for file in files:
            if file.lower() == "water_consolidated.xlsx":
                return os.path.join(root, file)
    return None

def main():
    try:
        # Azure setup
        blob_service = connect_to_azure()
        clients = [c.name for c in blob_service.list_containers()]
        
        # User selection
        print("Available clients:")
        for i, client in enumerate(clients, 1):
            print(f"{i}: {client}")
        selected_client = clients[int(input("Select client: "))-1]
        
        container = blob_service.get_container_client(selected_client)
        folders = sorted({os.path.dirname(b.name) for b in container.list_blobs() if b.name})
        
        print("Available folders:")
        for i, folder in enumerate(folders, 1):
            print(f"{i}: {folder}")
        selected_folder = folders[int(input("Select folder: "))-1]
        
        blobs = [b.name for b in container.list_blobs(name_starts_with=selected_folder)]
        print("Available files:")
        for i, blob in enumerate(blobs, 1):
            print(f"{i}: {blob}")
        selected = input("Select files (comma separated or 'all'): ")
        selected_blobs = blobs if selected.lower() == 'all' else [blobs[int(i)-1] for i in selected.split(',')]
        
        # Download files
        download_folder = input("Enter download folder path: ").strip()
        os.makedirs(download_folder, exist_ok=True)
        
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = []
            for blob in selected_blobs:
                dest = os.path.join(download_folder, os.path.relpath(blob, selected_folder))
                futures.append(executor.submit(download_blob, container, blob, dest))
            [f.result() for f in futures]
            
            # Unzip if needed
            for blob in selected_blobs:
                if blob.endswith('.zip'):
                    dest = os.path.join(download_folder, os.path.relpath(blob, selected_folder))
                    with zipfile.ZipFile(dest, 'r') as zip_ref:
                        zip_ref.extractall(os.path.dirname(dest))
        
        # Process files
        rename_files(download_folder, RENAME_MAPPING)
        remove_last_row_for_energy_files(download_folder)

        # Water file processing
        water_path = find_water_file(download_folder)
        if water_path:
            reporting_month = input("\nEnter reporting month (e.g., 'Feb-25'): ").strip()
            zero_out_cells_in_next_month(water_path, reporting_month)

        cleanup_temp_files(download_folder)
        display_files(download_folder)
        logging.info("Processing completed successfully!")
                
    except Exception as e:
        logging.error(f"Fatal error: {e}")
        raise

if __name__ == "__main__":
    main()